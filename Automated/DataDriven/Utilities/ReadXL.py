import openpyxl


class ReadXLClass:
    # constructor to create obj of ReadXL class
    def __init__(self, fileNamePath, sheetName):
        self.wk = openpyxl.load_workbook(fileNamePath)
        self.sh = self.wk[sheetName]

        self.row_count = self.sh.max_row
        self.col_count = self.sh.max_column

    # returns array with total count of rows, columns
    def fetch_total_row_col_counts(self):
        self.row_count = self.sh.max_row
        self.col_count = self.sh.max_column
        arr = [self.row_count, self.col_count]
        return arr

    # returns array with values of first row header
    def fetch_header_titles(self):
        keys = []
        for i in range(1, self.col_count + 1):
            cell = self.sh.cell(1, i).value
            keys.append(cell)
        return keys

    # returns array of values in one record
    def fetch_record_values(self, rowNumber):
        record = []
        for i in range(1, self.col_count + 1):
            cell = self.sh.cell(rowNumber, i).value
            record.append(cell)
        return record

    # returns json key-value of values in one record
    def fetch_record_json(self, rowNumber):
        keys = self.fetch_header_titles()
        record = self.fetch_record_values(rowNumber)
        json_body = {}

        for i in range(len(keys)):
            json_body[keys[i]] = record[i]

        return json_body
